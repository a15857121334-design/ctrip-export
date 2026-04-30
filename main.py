from __future__ import annotations

import argparse
import os
import re
import shutil
import subprocess
import sys
import time
from copy import copy, deepcopy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any


EXCEL_HEADERS = [
    "日期",
    "操作人",
    "来源",
    "目的地",
    "供应商",
    "团期",
    "人数",
    "卖价",
    "优惠后卖价",
    "结算",
    "加返后结算",
    "利润",
    "备注",
]

PLACEHOLDER_URL_PARTS = ("example.com", "replace-with", "TODO")


class ConfigError(RuntimeError):
    pass


@dataclass
class AppConfig:
    raw: dict[str, Any]
    config_path: Path
    template_path: Path
    output_path: Path
    storage_state_path: Path


def import_yaml():
    try:
        import yaml  # type: ignore
    except ImportError as exc:
        raise ConfigError(
            "缺少 PyYAML。请先运行：python -m pip install -r requirements.txt"
        ) from exc
    return yaml


def import_playwright():
    try:
        from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise ConfigError(
            "缺少 Playwright。请先运行：python -m pip install -r requirements.txt，"
            "然后运行：python -m playwright install chromium"
        ) from exc

    browsers_dir = Path(__file__).resolve().parent / ".ms-playwright"
    os.environ.setdefault("PLAYWRIGHT_BROWSERS_PATH", str(browsers_dir))

    def has_chromium_install() -> bool:
        chromium_ok = any(
            (path / "chrome-win64" / "chrome.exe").exists()
            for path in browsers_dir.glob("chromium-*")
        )
        headless_shell_ok = any(
            (path / "chrome-headless-shell-win64" / "chrome-headless-shell.exe").exists()
            for path in browsers_dir.glob("chromium_headless_shell-*")
        )
        return chromium_ok and headless_shell_ok

    if not has_chromium_install():
        print(f"检测到 Playwright Chromium 未安装或不完整，正在下载到：{browsers_dir}")
        env = os.environ.copy()
        env["PLAYWRIGHT_BROWSERS_PATH"] = str(browsers_dir)
        result = subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium"],
            env=env,
        )
        if result.returncode != 0 or not has_chromium_install():
            raise ConfigError(
                "Playwright 浏览器缺失且自动下载失败。请手动运行：python -m playwright install chromium"
            )

    return sync_playwright, PlaywrightTimeoutError


def resolve_path(value: str | None, base_dir: Path) -> Path:
    if not value:
        raise ConfigError("配置里的路径不能为空。")
    expanded = os.path.expandvars(os.path.expanduser(value))
    path = Path(expanded)
    if not path.is_absolute():
        path = base_dir / path
    return path.resolve()


def load_config(config_file: str | Path = "config.yaml") -> AppConfig:
    config_path = Path(config_file).resolve()
    if not config_path.exists():
        raise ConfigError(f"找不到配置文件：{config_path}")

    yaml = import_yaml()
    with config_path.open("r", encoding="utf-8") as handle:
        raw = yaml.safe_load(handle) or {}
    if not isinstance(raw, dict):
        raise ConfigError("config.yaml 必须是 YAML 字典结构。")

    base_dir = config_path.parent
    return AppConfig(
        raw=raw,
        config_path=config_path,
        template_path=resolve_path(raw.get("template_path"), base_dir),
        output_path=resolve_path(raw.get("output_path"), base_dir),
        storage_state_path=resolve_path(raw.get("storage_state_path"), base_dir),
    )


def is_placeholder_url(url: str | None) -> bool:
    if not url:
        return True
    return any(part.lower() in url.lower() for part in PLACEHOLDER_URL_PARTS)


def nested(config: dict[str, Any], *keys: str, default: Any = None) -> Any:
    current: Any = config
    for key in keys:
        if not isinstance(current, dict) or key not in current:
            return default
        current = current[key]
    return current


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def read_mapping(config: AppConfig) -> dict[str, dict[str, Any]]:
    mapping = config.raw.get("field_mapping") or {}
    if not isinstance(mapping, dict):
        raise ConfigError("field_mapping 必须是字典。")
    normalized: dict[str, dict[str, Any]] = {}
    for header in EXCEL_HEADERS:
        spec = mapping.get(header, {})
        if spec is None:
            spec = {}
        if not isinstance(spec, dict):
            raise ConfigError(f"field_mapping.{header} 必须是字典。")
        normalized[header] = spec
    return normalized


def check_config(config: AppConfig) -> int:
    errors: list[str] = []
    warnings: list[str] = []
    url = config.raw.get("order_page_url")
    selectors = config.raw.get("selectors") or {}
    field_mapping = read_mapping(config)

    if is_placeholder_url(url):
        errors.append("order_page_url 仍是占位地址，请改成携程后台订单管理页 URL。")
    if not config.template_path.exists():
        errors.append(f"找不到 Excel 模板：{config.template_path}")
    if not config.output_path.parent.exists():
        errors.append(f"输出目录不存在：{config.output_path.parent}")
    if not config.storage_state_path.exists():
        warnings.append(
            f"尚未保存登录态：{config.storage_state_path}。请先运行 save_login_state.py。"
        )

    required_selectors = ("table_rows",)
    for key in required_selectors:
        if not selectors.get(key):
            errors.append(f"selectors.{key} 不能为空。")

    date_selectors = ("start_date_input", "end_date_input", "search_button")
    missing_date = [key for key in date_selectors if not selectors.get(key)]
    if missing_date:
        warnings.append(
            "日期筛选选择器未填完整："
            + ", ".join(f"selectors.{key}" for key in missing_date)
            + "。脚本会跳过对应动作。"
        )

    for header, spec in field_mapping.items():
        if spec.get("type") == "formula":
            continue
        has_source = bool(spec.get("selector") or spec.get("header") or spec.get("source"))
        has_default = "default" in spec
        if not has_source and not has_default:
            warnings.append(f"字段 {header} 没有 selector/header/default，结果会为空。")
    if not selectors.get("table_headers"):
        header_only = [
            header
            for header, spec in field_mapping.items()
            if spec.get("header") and not spec.get("selector") and spec.get("type") != "formula"
        ]
        if header_only:
            warnings.append(
                "当前字段主要依赖 header 映射，但 selectors.table_headers 为空；"
                "请填写表头选择器，或为字段改用 selector。"
            )

    print("配置检查结果")
    print(f"- 配置文件：{config.config_path}")
    print(f"- 模板文件：{config.template_path}")
    print(f"- 输出文件：{config.output_path}")
    print(f"- 登录态文件：{config.storage_state_path}")
    if warnings:
        print("\n提醒：")
        for item in warnings:
            print(f"- {item}")
    if errors:
        print("\n需要处理：")
        for item in errors:
            print(f"- {item}")
        return 1
    print("\n配置基础检查通过。")
    return 0


def require_confirmation(message: str, accepted: set[str] | None = None) -> None:
    accepted_values = accepted or {"确认", "确认写入", "yes", "y"}
    print(message)
    answer = input("请输入“确认”继续，输入其他内容取消：").strip()
    if answer.lower() not in {value.lower() for value in accepted_values}:
        raise SystemExit("已取消，未写入任何文件。")


def wait_for_page_idle(page: Any, config: AppConfig, after_ms: int | None = None) -> None:
    timeout_ms = int(nested(config.raw, "browser", "timeout_ms", default=30000))
    loading_selector = nested(config.raw, "selectors", "loading_indicator", default="")
    try:
        page.wait_for_load_state("networkidle", timeout=timeout_ms)
    except Exception:
        pass
    if loading_selector:
        try:
            page.locator(loading_selector).first.wait_for(state="hidden", timeout=timeout_ms)
        except Exception:
            pass
    if after_ms:
        time.sleep(after_ms / 1000)


def fill_if_configured(page: Any, selector: str | None, value: str, label: str, timeout_ms: int) -> None:
    if not selector:
        print(f"- 跳过{label}：未配置选择器。")
        return
    locator = page.locator(selector).first
    locator.wait_for(state="visible", timeout=timeout_ms)
    try:
        locator.fill(value)
    except Exception:
        # Some Ant Design date inputs are readonly but still react to native
        # value/input/change events. This keeps date filtering configurable.
        locator.evaluate(
            """(el, value) => {
                const setter = Object.getOwnPropertyDescriptor(
                    window.HTMLInputElement.prototype,
                    'value'
                ).set;
                setter.call(el, value);
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
                el.dispatchEvent(new Event('blur', { bubbles: true }));
            }""",
            value,
        )


def set_date_if_configured(page: Any, selector: str | None, value: str, label: str, timeout_ms: int) -> None:
    if not selector:
        print(f"- 跳过{label}：未配置选择器。")
        return
    target = datetime.strptime(value, "%Y-%m-%d").date()
    locator = page.locator(selector).first
    locator.wait_for(state="visible", timeout=timeout_ms)
    locator.click()
    dropdown = page.locator(".ant-picker-dropdown:not(.ant-picker-dropdown-hidden)").last
    dropdown.wait_for(state="visible", timeout=timeout_ms)

    for _ in range(36):
        header_text = normalize_text(dropdown.locator(".ant-picker-header-view").first.inner_text())
        match = re.search(r"(\d{4})年\s*(\d{1,2})月", header_text)
        if not match:
            break
        current_year = int(match.group(1))
        current_month = int(match.group(2))
        current_key = current_year * 12 + current_month
        target_key = target.year * 12 + target.month
        if current_key == target_key:
            break
        button = ".ant-picker-header-next-btn" if current_key < target_key else ".ant-picker-header-prev-btn"
        dropdown.locator(button).first.click()
        page.wait_for_timeout(150)

    cell = dropdown.locator(f'td[title="{value}"]:not(.ant-picker-cell-disabled)').first
    cell.wait_for(state="visible", timeout=timeout_ms)
    cell.click()
    page.wait_for_timeout(300)


def click_if_configured(page: Any, selector: str | None, label: str, timeout_ms: int) -> bool:
    if not selector:
        print(f"- 跳过{label}：未配置选择器。")
        return False
    locator = page.locator(selector).first
    locator.wait_for(state="visible", timeout=timeout_ms)
    locator.click()
    return True


def exclude_order_statuses(page: Any, config: AppConfig) -> None:
    statuses = nested(config.raw, "filters", "exclude_order_statuses", default=[])
    if not statuses:
        return
    page.evaluate(
        """(statuses) => {
            const clean = (text) => (text || '')
                .replace(/\\s+/g, '')
                .replace(/[^\\u4e00-\\u9fa5]/g, '');
            const anchors = Array.from(document.querySelectorAll('a'));
            for (const status of statuses) {
                const target = anchors.find((a) => clean(a.innerText || a.textContent) === status);
                if (target && target.classList.contains('cur')) {
                    target.click();
                }
            }
        }""",
        statuses,
    )


def build_header_index(page: Any, config: AppConfig) -> dict[str, int]:
    selector = nested(config.raw, "selectors", "table_headers", default="")
    if not selector:
        return {}
    texts = page.locator(selector).all_inner_texts()
    header_index: dict[str, int] = {}
    for index, text in enumerate(texts):
        normalized = normalize_text(text)
        if normalized and normalized not in header_index:
            header_index[normalized] = index
    return header_index


def read_locator_text(locator: Any, spec: dict[str, Any], timeout_ms: int) -> str:
    attr = spec.get("attr") or "text"
    if attr == "text":
        return normalize_text(locator.inner_text(timeout=timeout_ms))
    if attr == "value":
        try:
            return normalize_text(locator.input_value(timeout=timeout_ms))
        except Exception:
            return normalize_text(locator.get_attribute("value", timeout=timeout_ms))
    return normalize_text(locator.get_attribute(attr, timeout=timeout_ms))


def apply_regex(value: str, spec: dict[str, Any]) -> str:
    pattern = spec.get("regex")
    if not pattern or not value:
        return value
    match = re.search(pattern, value)
    if not match:
        return ""
    if match.groups():
        return match.group(1)
    return match.group(0)


OPERATOR_ALIASES = {
    "闫硕奇": "闫",
    "闫": "闫",
    "迟丛丹": "迟",
    "迟": "迟",
}

COUNTRY_ALIASES = {
    "日本": ["日本", "东京", "大阪", "京都", "奈良", "北海道", "冲绳"],
    "韩国": ["韩国", "首尔", "釜山", "济州", "济州岛"],
    "澳大利亚": ["澳大利亚", "澳洲", "澳签"],
    "新西兰": ["新西兰"],
    "美国": ["美国", "美签"],
    "加拿大": ["加拿大", "加签"],
    "英国": ["英国", "英签"],
    "法国": ["法国", "巴黎"],
    "瑞士": ["瑞士"],
    "意大利": ["意大利", "罗马", "米兰", "威尼斯"],
    "德国": ["德国"],
    "西班牙": ["西班牙"],
    "葡萄牙": ["葡萄牙"],
    "希腊": ["希腊"],
    "土耳其": ["土耳其"],
    "埃及": ["埃及"],
    "阿联酋": ["阿联酋", "迪拜", "阿布扎比"],
    "泰国": ["泰国", "曼谷", "普吉", "清迈"],
    "新加坡": ["新加坡"],
    "马来西亚": ["马来西亚", "沙巴", "吉隆坡"],
    "印度尼西亚": ["印度尼西亚", "印尼", "巴厘岛"],
    "越南": ["越南", "芽庄", "岘港", "河内", "胡志明"],
    "柬埔寨": ["柬埔寨", "吴哥"],
    "菲律宾": ["菲律宾", "长滩", "薄荷"],
    "马尔代夫": ["马尔代夫"],
    "斯里兰卡": ["斯里兰卡"],
    "尼泊尔": ["尼泊尔"],
    "南非": ["南非"],
    "津巴布韦": ["津巴布韦"],
    "赞比亚": ["赞比亚"],
    "摩洛哥": ["摩洛哥"],
    "俄罗斯": ["俄罗斯", "莫斯科", "圣彼得堡"],
    "香港": ["香港"],
    "澳门": ["澳门"],
}

COUNTRY_COMPOUNDS = {
    "法瑞意": ["法国", "瑞士", "意大利"],
    "德法瑞意": ["德国", "法国", "瑞士", "意大利"],
    "法意瑞": ["法国", "意大利", "瑞士"],
    "英法": ["英国", "法国"],
}

FOREIGN_ISLAND_ALIASES = {
    "济州岛": ["济州岛", "济州"],
    "普吉岛": ["普吉岛", "普吉"],
    "苏梅岛": ["苏梅岛", "苏梅"],
    "巴厘岛": ["巴厘岛", "巴厘"],
    "长滩岛": ["长滩岛", "长滩"],
    "薄荷岛": ["薄荷岛", "薄荷"],
    "沙巴": ["沙巴"],
    "芽庄": ["芽庄"],
    "岘港": ["岘港"],
    "冲绳": ["冲绳"],
}

FOREIGN_ISLAND_COUNTRIES = {
    "济州岛": "韩国",
    "普吉岛": "泰国",
    "苏梅岛": "泰国",
    "巴厘岛": "印度尼西亚",
    "长滩岛": "菲律宾",
    "薄荷岛": "菲律宾",
    "沙巴": "马来西亚",
    "芽庄": "越南",
    "岘港": "越南",
    "冲绳": "日本",
}

DOMESTIC_CITIES = [
    "北京",
    "上海",
    "天津",
    "重庆",
    "杭州",
    "苏州",
    "南京",
    "无锡",
    "扬州",
    "嘉兴",
    "湖州",
    "宁波",
    "舟山",
    "绍兴",
    "金华",
    "台州",
    "温州",
    "丽水",
    "衢州",
    "广州",
    "深圳",
    "珠海",
    "佛山",
    "东莞",
    "成都",
    "西安",
    "武汉",
    "长沙",
    "厦门",
    "青岛",
    "昆明",
    "大理",
    "丽江",
    "西双版纳",
    "三亚",
    "海口",
    "桂林",
    "哈尔滨",
    "沈阳",
    "大连",
    "长春",
    "呼伦贝尔",
    "乌鲁木齐",
    "阿勒泰",
    "喀什",
    "拉萨",
    "林芝",
    "贵阳",
    "黄山",
    "合肥",
    "南昌",
    "福州",
    "泉州",
    "武夷山",
    "张家界",
    "恩施",
    "景德镇",
    "嵊泗",
    "嘉峪关",
    "张掖",
    "敦煌",
]

LANDMARK_CITY_ALIASES = {
    "乌镇": "嘉兴",
    "西塘": "嘉兴",
    "千岛湖": "杭州",
    "西湖": "杭州",
    "灵隐": "杭州",
    "普陀山": "舟山",
    "中山陵": "南京",
    "瘦西湖": "扬州",
    "寒山寺": "苏州",
    "长白山": "白山",
    "中国陶瓷博物馆": "景德镇",
    "陶阳里": "景德镇",
    "瑶里": "景德镇",
    "莫高窟": "敦煌",
    "青海湖": "海北",
    "茶卡盐湖": "海西",
}


def ordered_unique_matches(text: str, aliases: dict[str, list[str]] | dict[str, str]) -> list[str]:
    matches: list[tuple[int, str]] = []
    for canonical, alias_value in aliases.items():
        alias_list = alias_value if isinstance(alias_value, list) else [canonical, alias_value]
        for alias in alias_list:
            index = text.find(alias)
            if index >= 0:
                matches.append((index, canonical))
                break
    seen: set[str] = set()
    result: list[str] = []
    for _, value in sorted(matches, key=lambda item: item[0]):
        if value not in seen:
            result.append(value)
            seen.add(value)
    return result


def domestic_city_aliases() -> dict[str, list[str]]:
    aliases: dict[str, list[str]] = {city: [city] for city in DOMESTIC_CITIES}
    for alias, city in LANDMARK_CITY_ALIASES.items():
        aliases.setdefault(city, [city]).append(alias)
    return aliases


def format_island_destinations(islands: list[str], countries: list[str]) -> str:
    island_destinations: list[str] = []
    for island in islands:
        default_country = FOREIGN_ISLAND_COUNTRIES.get(island, "")
        country = default_country if default_country in countries or default_country else ""
        if not country and len(countries) == 1:
            country = countries[0]
        island_destinations.append(f"{country}{island}" if country else island)
    return "+".join(dict.fromkeys(island_destinations))


def normalize_operator(value: Any) -> str:
    text = normalize_text(value)
    return OPERATOR_ALIASES.get(text, text[:1])


def normalize_supplier(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    clean = text.replace("供应商：", "").strip()
    self_terms = ["携程贴牌自营", "贴牌自营", "携程国旅自营", "携程自营"]
    has_self = any(term in clean for term in self_terms)
    has_label = "贴牌" in clean
    remainder = clean
    for term in self_terms + ["携程", "自营", "：", ":", "+"]:
        remainder = remainder.replace(term, "")
    remainder = remainder.strip(" -_/（）()，,")
    if has_label or (has_self and remainder):
        return f"贴牌自营+{remainder}" if remainder else "贴牌自营"
    if has_self:
        return "携程自营"
    return clean


def normalize_destination(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    compact = re.sub(r"\s+", "", text)

    countries: list[str] = []
    for marker, marker_countries in COUNTRY_COMPOUNDS.items():
        if marker in compact:
            countries.extend(marker_countries)
    countries.extend(ordered_unique_matches(compact, COUNTRY_ALIASES))
    countries = list(dict.fromkeys(countries))

    is_visa = any(word in compact for word in ["签证", "签注", "签"])
    if is_visa:
        if not countries and "澳" in compact:
            countries = ["澳大利亚"]
        if countries:
            return "+".join(f"{country}签证" for country in countries)

    cities = ordered_unique_matches(compact, domestic_city_aliases())
    if "定制" in compact and cities:
        return f"{'+'.join(cities)}定制"

    islands = ordered_unique_matches(compact, FOREIGN_ISLAND_ALIASES)
    if "自由行" in compact:
        if islands:
            return f"{format_island_destinations(islands, countries)}自由行"
        if countries:
            return f"{'+'.join(countries)}自由行"
        if cities:
            return f"{'+'.join(cities)}自由行"

    if islands:
        return format_island_destinations(islands, countries)

    if countries:
        return "+".join(countries)

    if cities:
        return "+".join(cities)

    fallback = re.split(r"[·*/（(]", compact, maxsplit=1)[0]
    if "-" in fallback:
        parts = [part for part in fallback.split("-") if part]
        if parts and all(part == parts[0] for part in parts):
            fallback = parts[0]
    fallback = re.sub(r"\d+[日天晚].*$", "", fallback)
    return fallback[:30]


def parse_people_count(value: Any) -> int | str | None:
    text = normalize_text(value)
    if not text:
        return None
    matches = re.findall(r"(\d+)\s*(?:成人|儿童|小孩|婴儿|人)", text)
    if matches:
        return sum(int(item) for item in matches)
    return parse_number(text, integer=True)


def apply_transform(value: Any, spec: dict[str, Any]) -> Any:
    transform = spec.get("transform")
    if transform == "operator_short":
        return normalize_operator(value)
    if transform == "destination":
        return normalize_destination(value)
    if transform == "supplier":
        return normalize_supplier(value)
    return value


def extract_order_metadata(row: Any, timeout_ms: int) -> dict[str, str]:
    text = normalize_text(row.inner_text(timeout=timeout_ms))
    order_no_match = re.search(r"订单：\s*(\d+)", text)
    status_match = re.search(r"添加备忘\s+(.+?)\s+查看", text)
    if not status_match:
        status_match = re.search(r"(待确认|已确认|已完成|已取消|全部退订|已归档)", text)
    return {
        "__order_no": order_no_match.group(1) if order_no_match else "",
        "__order_status": status_match.group(1).strip() if status_match else "",
    }


def dedupe_value(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if value is None:
        return ""
    return normalize_text(value)


def duplicate_key_value(field: str, value: Any) -> str:
    if field in {"日期", "团期"}:
        parsed = parse_date(value)
        return parsed.isoformat() if isinstance(parsed, date) else dedupe_value(parsed)
    if field in {"人数"}:
        parsed = parse_people_count(value)
        return str(parsed or "")
    if field in {"卖价", "优惠后卖价", "结算", "加返后结算"}:
        parsed = parse_number(value)
        if isinstance(parsed, (int, float)):
            return f"{parsed:.2f}"
        return dedupe_value(parsed)
    if field == "供应商":
        supplier = normalize_supplier(value)
        return supplier.removeprefix("贴牌自营+")
    return dedupe_value(value)


def duplicate_key(order: dict[str, Any], key_fields: list[str]) -> tuple[str, ...]:
    return tuple(duplicate_key_value(field, order.get(field)) for field in key_fields)


def dedupe_score(order: dict[str, Any], config: AppConfig) -> tuple[int, float, float]:
    status = normalize_text(order.get("__order_status"))
    sell_price = parse_number(order.get("卖价")) or 0
    settlement = parse_number(order.get("结算")) or 0
    if not isinstance(sell_price, (int, float)):
        sell_price = 0
    if not isinstance(settlement, (int, float)):
        settlement = 0

    prefer_non_cancelled = bool(nested(config.raw, "dedupe", "prefer_non_cancelled", default=True))
    active_score = 0
    if prefer_non_cancelled:
        active_score = 0 if any(word in status for word in ["取消", "退订"]) else 1
    return active_score, float(sell_price), float(settlement)


def dedupe_orders(orders: list[dict[str, Any]], config: AppConfig) -> list[dict[str, Any]]:
    if not bool(nested(config.raw, "dedupe", "enabled", default=True)):
        return orders
    key_fields = nested(
        config.raw,
        "dedupe",
        "key_fields",
        default=["日期", "操作人", "目的地", "供应商", "团期", "人数"],
    )
    if not isinstance(key_fields, list) or not key_fields:
        return orders

    selected: dict[tuple[str, ...], dict[str, Any]] = {}
    order_keys: list[tuple[str, ...]] = []
    removed = 0
    for order in orders:
        key = duplicate_key(order, key_fields)
        if key not in selected:
            selected[key] = order
            order_keys.append(key)
            continue
        removed += 1
        if dedupe_score(order, config) > dedupe_score(selected[key], config):
            selected[key] = order

    if removed:
        print(f"- 已按业务字段去重，剔除重复订单 {removed} 行。")
    return [selected[key] for key in order_keys]


def template_existing_orders(config: AppConfig) -> list[dict[str, Any]]:
    if not config.template_path.exists():
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    workbook = load_workbook(config.template_path, data_only=False)
    sheet_name = nested(config.raw, "excel", "template_sheet_name", default="工作表1")
    if sheet_name not in workbook.sheetnames:
        return []
    ws = workbook[sheet_name]
    existing: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        order = {header: ws.cell(row, col).value for col, header in enumerate(EXCEL_HEADERS, start=1)}
        non_formula_values = [
            value
            for header, value in order.items()
            if header != "利润" and value not in ("", None)
        ]
        if non_formula_values:
            existing.append(order)
    return existing


def skip_existing_template_orders(orders: list[dict[str, Any]], config: AppConfig) -> list[dict[str, Any]]:
    if not bool(nested(config.raw, "dedupe", "skip_existing_template", default=True)):
        return orders
    key_fields = nested(
        config.raw,
        "dedupe",
        "existing_key_fields",
        default=["日期", "操作人", "供应商", "团期", "人数", "卖价"],
    )
    if not isinstance(key_fields, list) or not key_fields:
        return orders

    existing_keys = {
        duplicate_key(order, key_fields)
        for order in template_existing_orders(config)
        if any(order.get(field) not in ("", None) for field in key_fields)
    }
    if not existing_keys:
        return orders

    kept: list[dict[str, Any]] = []
    skipped = 0
    for order in orders:
        if duplicate_key(order, key_fields) in existing_keys:
            skipped += 1
        else:
            kept.append(order)
    if skipped:
        print(f"- 已跳过模板中已有订单 {skipped} 行。")
    return kept


def extract_field_from_row(
    row: Any,
    header: str,
    spec: dict[str, Any],
    header_index: dict[str, int],
    config: AppConfig,
) -> Any:
    timeout_ms = int(nested(config.raw, "browser", "timeout_ms", default=30000))
    row_cells_selector = nested(config.raw, "selectors", "row_cells", default="td, th")
    raw_value = ""
    source = spec.get("source")

    selector = spec.get("selector")
    if source == "row_text":
        raw_value = normalize_text(row.inner_text(timeout=timeout_ms))
    elif selector:
        locator = row.locator(selector).first
        if locator.count() > 0:
            raw_value = read_locator_text(locator, spec, timeout_ms)
    else:
        web_header = normalize_text(spec.get("header") or header)
        cell_index = header_index.get(web_header)
        if cell_index is not None:
            cells = row.locator(row_cells_selector)
            if cells.count() > cell_index:
                raw_value = read_locator_text(cells.nth(cell_index), spec, timeout_ms)

    raw_value = apply_regex(raw_value, spec)
    if raw_value == "":
        raw_value = spec.get("default", "")
    raw_value = apply_transform(raw_value, spec)
    return convert_value(raw_value, spec.get("type", "text"))


def parse_date(value: Any) -> date | str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = normalize_text(value)
    text = re.sub(r"[年月.]", "-", text).replace("日", "")
    text = text.split(" ")[0]
    text = text.split("T")[0]
    formats = ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d", "%m-%d-%Y", "%m/%d/%Y")
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return normalize_text(value)


def parse_number(value: Any, integer: bool = False) -> int | float | str | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value) if integer else value

    text = normalize_text(value)
    negative = text.startswith("(") and text.endswith(")")
    cleaned = re.sub(r"[,\s￥¥元人]", "", text)
    cleaned = cleaned.strip("()")
    match = re.search(r"-?\d+(?:\.\d+)?", cleaned)
    if not match:
        return text
    number = float(match.group(0))
    if negative:
        number = -number
    if integer:
        return int(round(number))
    return int(number) if number.is_integer() else number


def convert_value(value: Any, value_type: str) -> Any:
    if value_type == "date":
        return parse_date(value)
    if value_type == "int":
        return parse_number(value, integer=True)
    if value_type == "people_count":
        return parse_people_count(value)
    if value_type == "number":
        return parse_number(value)
    if value_type == "formula":
        return None
    return normalize_text(value)


def scrape_orders(config: AppConfig, start_date: str, end_date: str) -> list[dict[str, Any]]:
    url = config.raw.get("order_page_url")
    if is_placeholder_url(url):
        raise ConfigError("请先在 config.yaml 中填写真实的 order_page_url。")
    if not config.storage_state_path.exists():
        raise ConfigError(
            f"找不到登录态文件：{config.storage_state_path}。请先运行 save_login_state.py。"
        )

    selectors = config.raw.get("selectors") or {}
    if not selectors.get("table_rows"):
        raise ConfigError("请先在 config.yaml 中填写 selectors.table_rows。")

    sync_playwright, _ = import_playwright()
    timeout_ms = int(nested(config.raw, "browser", "timeout_ms", default=30000))
    headless = bool(nested(config.raw, "browser", "headless", default=False))
    slow_mo_ms = int(nested(config.raw, "browser", "slow_mo_ms", default=0))
    max_pages = int(nested(config.raw, "scraping", "max_pages", default=100))
    wait_after_search_ms = int(nested(config.raw, "scraping", "wait_after_search_ms", default=1000))
    wait_after_page_ms = int(nested(config.raw, "scraping", "wait_after_page_ms", default=800))
    mapping = read_mapping(config)

    orders: list[dict[str, Any]] = []
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=headless, slow_mo=slow_mo_ms)
        context = browser.new_context(storage_state=str(config.storage_state_path))
        page = context.new_page()
        page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
        wait_for_page_idle(page, config)

        set_date_if_configured(page, selectors.get("start_date_input"), start_date, "开始日期", timeout_ms)
        set_date_if_configured(page, selectors.get("end_date_input"), end_date, "结束日期", timeout_ms)
        exclude_order_statuses(page, config)
        clicked = click_if_configured(page, selectors.get("search_button"), "查询按钮", timeout_ms)
        wait_for_page_idle(page, config, wait_after_search_ms if clicked else None)

        for page_number in range(1, max_pages + 1):
            header_index = build_header_index(page, config)
            row_locator = page.locator(selectors["table_rows"])
            try:
                row_locator.first.wait_for(state="attached", timeout=timeout_ms)
            except Exception:
                pass
            row_count = row_locator.count()
            print(f"- 第 {page_number} 页读取到 {row_count} 行。")

            for row_index in range(row_count):
                row = row_locator.nth(row_index)
                if not normalize_text(row.inner_text(timeout=timeout_ms)):
                    continue
                order: dict[str, Any] = {}
                for header in EXCEL_HEADERS:
                    spec = mapping.get(header, {})
                    if spec.get("type") == "formula" or header == "利润":
                        continue
                    order[header] = extract_field_from_row(row, header, spec, header_index, config)
                if any(value not in ("", None) for value in order.values()):
                    order.update(extract_order_metadata(row, timeout_ms))
                    excluded_statuses = nested(config.raw, "filters", "exclude_order_statuses", default=[])
                    if order.get("__order_status") in excluded_statuses:
                        continue
                    orders.append(order)

            next_selector = selectors.get("next_page_button")
            if not next_selector:
                break
            next_button = page.locator(next_selector).first
            if next_button.count() == 0 or next_button_disabled(next_button, config):
                break
            next_button.click()
            wait_for_page_idle(page, config, wait_after_page_ms)
        else:
            print(f"- 已达到最大翻页数 {max_pages}，停止翻页。")

        context.close()
        browser.close()
    unique_orders = dedupe_orders(orders, config)
    return sort_orders_by_order_date(skip_existing_template_orders(unique_orders, config))


def next_button_disabled(button: Any, config: AppConfig) -> bool:
    attr_name = nested(config.raw, "selectors", "next_page_disabled_attribute", default="disabled")
    disabled_class = nested(config.raw, "selectors", "next_page_disabled_class", default="disabled")
    try:
        if button.is_disabled():
            return True
    except Exception:
        pass
    try:
        if attr_name:
            attr_value = button.get_attribute(attr_name)
            if attr_value is not None and attr_value.lower() not in {"", "false", "0"}:
                return True
    except Exception:
        pass
    try:
        if button.get_attribute("aria-disabled") == "true":
            return True
    except Exception:
        pass
    try:
        class_name = button.get_attribute("class") or ""
        if disabled_class and disabled_class in class_name.split():
            return True
    except Exception:
        pass
    return False


def copy_row_format(source_ws: Any, source_row: int, target_ws: Any, target_row: int, max_col: int) -> None:
    target_ws.row_dimensions[target_row].height = source_ws.row_dimensions[source_row].height
    target_ws.row_dimensions[target_row].hidden = source_ws.row_dimensions[source_row].hidden
    for col in range(1, max_col + 1):
        source_cell = source_ws.cell(source_row, col)
        target_cell = target_ws.cell(target_row, col)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)


def prepare_result_sheet(workbook: Any, config: AppConfig) -> Any:
    template_sheet_name = nested(config.raw, "excel", "template_sheet_name", default="工作表1")
    result_sheet_name = config.raw.get("result_sheet_name") or "ctrip"
    if template_sheet_name not in workbook.sheetnames:
        raise ConfigError(f"模板中找不到工作表：{template_sheet_name}")

    template_ws = workbook[template_sheet_name]
    if result_sheet_name in workbook.sheetnames:
        del workbook[result_sheet_name]

    result_ws = workbook.copy_worksheet(template_ws)
    result_ws.title = result_sheet_name
    result_ws.freeze_panes = template_ws.freeze_panes
    result_ws.auto_filter.ref = template_ws.auto_filter.ref
    try:
        result_ws.data_validations = deepcopy(template_ws.data_validations)
    except Exception:
        pass
    try:
        result_ws.conditional_formatting = deepcopy(template_ws.conditional_formatting)
    except Exception:
        pass
    return result_ws


def write_excel(config: AppConfig, orders: list[dict[str, Any]]) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ConfigError(
            "缺少 openpyxl。请先运行：python -m pip install -r requirements.txt"
        ) from exc

    if not config.template_path.exists():
        raise ConfigError(f"找不到模板文件：{config.template_path}")

    workbook = load_workbook(config.template_path)
    result_ws = prepare_result_sheet(workbook, config)
    template_ws = workbook[nested(config.raw, "excel", "template_sheet_name", default="工作表1")]
    max_col = len(EXCEL_HEADERS)
    end_row = max(result_ws.max_row, len(orders) + 1)

    for col, header in enumerate(EXCEL_HEADERS, start=1):
        result_ws.cell(1, col).value = header

    for row in range(2, end_row + 1):
        for col in range(1, max_col + 1):
            result_ws.cell(row, col).value = None

    style_source_row = int(nested(config.raw, "excel", "style_source_row", default=20))
    if style_source_row > template_ws.max_row:
        style_source_row = 2

    for offset, order in enumerate(orders, start=2):
        copy_row_format(template_ws, style_source_row, result_ws, offset, max_col)
        for col, header in enumerate(EXCEL_HEADERS, start=1):
            cell = result_ws.cell(offset, col)
            if header == "利润":
                cell.value = f'=IF(I{offset}="","",I{offset}-K{offset})'
            else:
                value = order.get(header)
                cell.value = "" if value is None else value

    result_ws.auto_filter.ref = f"A1:M{max(1, len(orders) + 1)}"
    config.output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(config.output_path)


def daily_backup_path(config: AppConfig) -> Path:
    backup_dir = config.template_path.parent / "ctrip_backup"
    return backup_dir / f"{config.template_path.stem}_{date.today().isoformat()}{config.template_path.suffix}"


def create_daily_backup(config: AppConfig) -> tuple[Path, bool]:
    if not config.template_path.exists():
        raise ConfigError(f"找不到订单文件：{config.template_path}")
    backup_path = daily_backup_path(config)
    if backup_path.exists():
        return backup_path, False
    backup_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(config.template_path, backup_path)
    return backup_path, True


def row_has_order_data(ws: Any, row: int) -> bool:
    for col, header in enumerate(EXCEL_HEADERS, start=1):
        if header == "利润":
            continue
        if ws.cell(row, col).value not in ("", None):
            return True
    return False


def last_order_row(ws: Any) -> int:
    for row in range(ws.max_row, 1, -1):
        if row_has_order_data(ws, row):
            return row
    return 1


def filter_end_row(ws: Any) -> int:
    ref = ws.auto_filter.ref or ""
    match = re.search(r":\$?[A-Z]+\$?(\d+)$", ref)
    if match:
        return int(match.group(1))
    return 1


def order_date_sort_key(value: Any) -> tuple[int, str]:
    parsed = parse_date(value)
    if isinstance(parsed, date):
        return 0, parsed.isoformat()
    return 1, dedupe_value(value)


def sort_orders_by_order_date(orders: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        order
        for _, order in sorted(
            enumerate(orders),
            key=lambda item: (
                order_date_sort_key(item[1].get("日期")),
                dedupe_value(item[1].get("__order_no")),
                item[0],
            ),
        )
    ]


def sort_sheet_rows_by_order_date(ws: Any, start_row: int = 2, end_row: int | None = None) -> None:
    if end_row is None:
        end_row = last_order_row(ws)
    if end_row < start_row:
        return

    rows: list[tuple[int, list[Any]]] = []
    for row in range(start_row, end_row + 1):
        rows.append((row, [ws.cell(row, col).value for col in range(1, len(EXCEL_HEADERS) + 1)]))

    sorted_rows = sorted(
        rows,
        key=lambda item: (order_date_sort_key(item[1][0]), item[0]),
    )
    for target_row, (_, values) in zip(range(start_row, end_row + 1), sorted_rows):
        for col, value in enumerate(values, start=1):
            ws.cell(target_row, col).value = value
        ws.cell(target_row, 12).value = f'=IF(I{target_row}="","",I{target_row}-K{target_row})'


def first_date_format(ws: Any, column: int, start_row: int, end_row: int, fallback: str) -> str:
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row, column)
        if isinstance(cell.value, (datetime, date)) and cell.number_format not in ("", "General"):
            return cell.number_format
    return fallback


def normalize_date_column_formats(ws: Any, start_row: int = 2, end_row: int | None = None) -> None:
    if end_row is None:
        end_row = last_order_row(ws)
    if end_row < start_row:
        return
    order_date_format = first_date_format(ws, 1, start_row, end_row, "yyyy-mm-dd")
    departure_date_format = first_date_format(ws, 6, start_row, end_row, "yyyy/m/d;@")
    for row in range(start_row, end_row + 1):
        if isinstance(ws.cell(row, 1).value, (datetime, date)):
            ws.cell(row, 1).number_format = order_date_format
        if isinstance(ws.cell(row, 6).value, (datetime, date)):
            ws.cell(row, 6).number_format = departure_date_format


def append_orders_to_template(config: AppConfig, orders: list[dict[str, Any]]) -> tuple[int, int]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ConfigError(
            "缺少 openpyxl。请先运行：python -m pip install -r requirements.txt"
        ) from exc

    if not config.template_path.exists():
        raise ConfigError(f"找不到订单文件：{config.template_path}")

    workbook = load_workbook(config.template_path)
    sheet_name = nested(config.raw, "excel", "template_sheet_name", default="工作表1")
    if sheet_name not in workbook.sheetnames:
        raise ConfigError(f"订单文件中找不到工作表：{sheet_name}")

    ws = workbook[sheet_name]
    for col, header in enumerate(EXCEL_HEADERS, start=1):
        if normalize_text(ws.cell(1, col).value) != header:
            raise ConfigError(f"订单表第 {col} 列表头不是 {header}，为避免写错列，已停止。")

    last_row = last_order_row(ws)
    start_row = last_row + 1
    style_source_row = int(nested(config.raw, "excel", "style_source_row", default=20))
    if style_source_row > ws.max_row or style_source_row < 2:
        style_source_row = max(2, last_row)

    if not orders:
        return start_row, start_row - 1

    sorted_orders = sort_orders_by_order_date(orders)
    max_col = len(EXCEL_HEADERS)
    for offset, order in enumerate(sorted_orders):
        row = start_row + offset
        copy_row_format(ws, style_source_row, ws, row, max_col)
        for col, header in enumerate(EXCEL_HEADERS, start=1):
            cell = ws.cell(row, col)
            if header == "利润":
                cell.value = f'=IF(I{row}="","",I{row}-K{row})'
            else:
                value = order.get(header)
                cell.value = "" if value is None else value

    end_row = start_row + len(orders) - 1
    sort_sheet_rows_by_order_date(ws, 2, end_row)
    normalize_date_column_formats(ws, 2, end_row)
    ws.auto_filter.ref = f"A1:M{max(filter_end_row(ws), end_row)}"
    try:
        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    except Exception:
        pass
    workbook.save(config.template_path)
    return start_row, start_row + len(orders) - 1


def print_preview(orders: list[dict[str, Any]], limit: int = 5) -> None:
    print(f"抓取预览：共 {len(orders)} 行。")
    for index, order in enumerate(orders[:limit], start=1):
        values = [order.get(header, "") for header in EXCEL_HEADERS if header != "利润"]
        print(f"{index}. {values}")
    if len(orders) > limit:
        print(f"... 还有 {len(orders) - limit} 行未显示。")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="从携程后台抓取订单并生成本地 Excel。")
    parser.add_argument("--config", default="config.yaml", help="配置文件路径。")
    parser.add_argument("--start-date", help="开始日期，格式 YYYY-MM-DD。")
    parser.add_argument("--end-date", help="结束日期，格式 YYYY-MM-DD。")
    parser.add_argument("--recent-days", type=int, help="查询最近 N 天订单，含今天。")
    parser.add_argument("--dry-run", action="store_true", help="只抓取并打印预览，不生成 Excel。")
    parser.add_argument("--check-config", action="store_true", help="只检查配置，不抓取。")
    parser.add_argument("--update-template", action="store_true", help="把新增订单直接追加到订单.xlsx。")
    parser.add_argument("--daily-backup", action="store_true", help="更新前按天备份一次订单.xlsx。")
    parser.add_argument("--write-preview", action="store_true", help="同时生成 ctrip.xlsx，内容为本次新增订单。")
    parser.add_argument("--yes", action="store_true", help="跳过写入确认，供定时任务使用。")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        config = load_config(args.config)
        if args.check_config:
            return check_config(config)

        start_date = args.start_date
        end_date = args.end_date
        if args.recent_days:
            if args.recent_days < 1:
                parser.error("--recent-days 必须大于等于 1。")
            end = date.today()
            start = end - timedelta(days=args.recent_days - 1)
            start_date = start.isoformat()
            end_date = end.isoformat()

        if not start_date or not end_date:
            parser.error("请同时提供 --start-date 和 --end-date，格式 YYYY-MM-DD。")

        orders = scrape_orders(config, start_date, end_date)
        print_preview(orders)
        if args.dry_run:
            print("dry-run 模式：未生成 Excel。")
            return 0

        if args.update_template:
            if not args.yes:
                require_confirmation(
                    "\n即将自动更新订单表：\n"
                    f"- 订单文件：{config.template_path}\n"
                    f"- 查询日期：{start_date} 至 {end_date}\n"
                    f"- 新增订单行数：{len(orders)}\n"
                    f"- 每日备份：{'开启' if args.daily_backup else '关闭'}\n"
                    f"- 新增数据预览文件：{config.output_path if args.write_preview else '不生成'}\n"
                    "将只追加去重后的新增订单，不会删除已有行。"
                )
            if args.daily_backup:
                backup_path, created = create_daily_backup(config)
                if created:
                    print(f"已创建今日备份：{backup_path}")
                else:
                    print(f"今日备份已存在，未覆盖：{backup_path}")
            start_row, end_row = append_orders_to_template(config, orders)
            if orders:
                print(f"已更新订单文件：{config.template_path}，新增行 {start_row}-{end_row}。")
            else:
                print(f"没有新增订单，订单文件未新增行：{config.template_path}")
            if args.write_preview:
                write_excel(config, orders)
                print(f"已生成本次新增数据预览：{config.output_path}")
            return 0

        exists_text = "会覆盖已有文件" if config.output_path.exists() else "将创建新文件"
        if not args.yes:
            require_confirmation(
                "\n即将写入 Excel：\n"
                f"- 只读模板：{config.template_path}\n"
                f"- 输出文件：{config.output_path}\n"
                f"- 写入订单行数：{len(orders)}\n"
                f"- 状态：{exists_text}\n"
                "原始模板不会被修改。"
            )
        write_excel(config, orders)
        print(f"已生成：{config.output_path}")
        return 0
    except ConfigError as exc:
        print(f"错误：{exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
